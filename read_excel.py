import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

try:
    wb = openpyxl.load_workbook('d:/방과후 징수 엑셀/방과후 기초자료.xlsx', data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f"=== Sheet: {sheet} (Rows: {ws.max_row}, Cols: {ws.max_column}) ===")
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            print(row)
        print("\n")
except Exception as e:
    print(f"Error: {e}")
